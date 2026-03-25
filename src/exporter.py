"""
Módulo de exportação de relatórios em Excel, PDF e CSV.
"""

import os
import csv
from datetime import datetime

EXPORTS_DIR = "exports"
os.makedirs(EXPORTS_DIR, exist_ok=True)


def _timestamp() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def exportar_excel(monitorados: dict) -> str:
    """Exporta as proposições monitoradas para Excel (.xlsx)."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Proposições Monitoradas"

    # Cabeçalho
    cabecalho = ["Casa", "Número", "Ano", "Ementa", "Autor", "Situação"]
    for col, titulo in enumerate(cabecalho, 1):
        cell = ws.cell(row=1, column=col, value=titulo)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1a3a5c")
        cell.alignment = Alignment(horizontal="center")

    # Dados
    for linha, (chave, dados) in enumerate(monitorados.items(), 2):
        casa, numero, ano = chave.split(":")
        ws.cell(row=linha, column=1, value=casa)
        ws.cell(row=linha, column=2, value=numero)
        ws.cell(row=linha, column=3, value=ano)
        ws.cell(row=linha, column=4, value=dados.get("ementa", ""))
        ws.cell(row=linha, column=5, value=dados.get("autor", ""))
        ws.cell(row=linha, column=6, value=dados.get("situacao", ""))

    # Ajusta largura das colunas
    larguras = [12, 10, 8, 60, 30, 35]
    for col, largura in enumerate(larguras, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = largura

    caminho = f"{EXPORTS_DIR}/proposicoes_{_timestamp()}.xlsx"
    wb.save(caminho)
    return caminho


def exportar_pdf(monitorados: dict) -> str:
    """Exporta as proposições monitoradas para PDF."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    caminho = f"{EXPORTS_DIR}/proposicoes_{_timestamp()}.pdf"
    doc = SimpleDocTemplate(caminho, pagesize=A4,
                            leftMargin=2*cm, rightMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)

    styles = getSampleStyleSheet()
    elementos = []

    # Título
    titulo_style = ParagraphStyle("titulo", parent=styles["Title"],
                                  fontSize=16, textColor=colors.HexColor("#1a3a5c"))
    elementos.append(Paragraph("Monitor Legislativo — Proposições Monitoradas", titulo_style))
    elementos.append(Spacer(1, 0.4*cm))
    elementos.append(Paragraph(
        f"Gerado em {datetime.now().strftime('%d/%m/%Y às %H:%M')}",
        styles["Normal"]
    ))
    elementos.append(Spacer(1, 0.6*cm))

    # Tabela
    dados_tabela = [["Casa", "PL", "Situação", "Ementa"]]
    for chave, dados in monitorados.items():
        casa, numero, ano = chave.split(":")
        ementa = dados.get("ementa", "")
        if len(ementa) > 80:
            ementa = ementa[:80] + "..."
        dados_tabela.append([
            casa,
            f"{numero}/{ano}",
            dados.get("situacao", "N/D"),
            ementa,
        ])

    tabela = Table(dados_tabela, colWidths=[3*cm, 2*cm, 5*cm, 7.5*cm])
    tabela.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a3a5c")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f0f4f8")]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cccccc")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("PADDING", (0, 0), (-1, -1), 4),
    ]))

    elementos.append(tabela)
    doc.build(elementos)
    return caminho


def exportar_csv(monitorados: dict) -> str:
    """Exporta as proposições monitoradas para CSV."""
    caminho = f"{EXPORTS_DIR}/proposicoes_{_timestamp()}.csv"

    with open(caminho, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=["Casa", "Número", "Ano", "Ementa", "Autor", "Situação"])
        writer.writeheader()
        for chave, dados in monitorados.items():
            casa, numero, ano = chave.split(":")
            writer.writerow({
                "Casa": casa,
                "Número": numero,
                "Ano": ano,
                "Ementa": dados.get("ementa", ""),
                "Autor": dados.get("autor", ""),
                "Situação": dados.get("situacao", ""),
            })

    return caminho
